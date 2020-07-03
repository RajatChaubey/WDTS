#!/usr/bin/python

import openpyxl
from openpyxl import load_workbook

book = openpyxl.load_workbook('FM_SNMP_Trap_OID_NoSeverity.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value) 
print(a3.value)
