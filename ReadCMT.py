#!/usr/bin/python
import subprocess
import openpyxl;
from openpyxl import load_workbook


out = subprocess.Popen('ls -lrt *CMTBuilder*', shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)

stdout,stderr = out.communicate()
print "This is CMT EXL:",(stdout.split()[8])


def CMTvalues():
    #book = openpyxl.load_workbook('fm_vmwr_vcentre_esxi_snmp_nr6.6-vmware-mib_CMTBuilder_v3.4.xlsx')
    book = openpyxl.load_workbook(stdout.split()[8])

    #print(book.get_sheet_names())

    active_sheet = book.active
    #print(type(active_sheet))

    sheet = book.get_sheet_by_name("CATEGORY")
    #print(sheet.title)

    a1 = sheet['B2']
    a2 = sheet['C2']
#a3 = sheet.cell(row=3, column=1)

    print "Match String value in CMTBuilder:",(a1.value)
    print "Plugin Name to be Tested:",(a2.value) 
#print(a3.value)
    Match_OID = a1.value.replace("\\","")
#print "This OID to be use in Sanity Scrpit: ",a1.value.replace("\\", ""),"\n"
    print "This OID to be use in Sanity Scrpit: ",Match_OID,"\n"
    sheet = book.get_sheet_by_name("ALARMS")
    #print(sheet.max_row)
    active_sheet = book.active
    m_row = sheet.max_row
    for i in range(1, m_row +2): 
              cell_obj = sheet.cell(row = i, column = 1)
          
              #print "This is to TEST VALUE of i :  ", i 
              if(cell_obj.value):
                           print "This is the ALARM ID/OID: ", (cell_obj.value)
   
              else:
                  Alarm_count = i-2
                  print "This is Alarm count:",i-2
                  break  

    return (Match_OID,Alarm_count,a2.value);

#OID,Alarm_count,Plugin_Name = CMTvalues();
#print "\n Required Values of CMT are:",OID,Alarm_count,Plugin_Name




def Class_values():
    
    book = openpyxl.load_workbook(stdout.split()[8])
    sheet = book.get_sheet_by_name("CLASSES")
    
    
    active_sheet = book.active
    m_row = sheet.max_row
    #C_List=[]
    line=""
    #print "Max row :",m_row
    for i in range(1, m_row +2):     # this should be m_row +1 ,for now its 2 ,but have to suggest developer that there should be no border in any of the sheet of CMT workbook
              cell_obj = sheet.cell(row = i, column = 1)
              if(cell_obj.value):
                           print "This is the Class Name: ", (cell_obj.value) 
                           line= line+' '+cell_obj.value                     
                           #C_List.append((cell_obj.value))
                           #print "List content are INSIDE LOOP:", C_List;
                           #print "Line content are INSIDE LOOP:", line
              else:
                  unit = i-2
                  print "This is count of CLASS in CMT:",unit
                  break
    print "Class n it's native classes  are:", line
    return line

#line=Class_values();
#print "List content are:",C_List
