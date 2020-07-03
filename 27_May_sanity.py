#!/usr/bin/python 
import os,shutil;
import cx_Oracle;
import openpyxl;
import sys;
import re;
import ReadCMT;

#ot=os.system('ls -lrt /opt/home/netx/backups/pa | tail -n 1');

#os.system('y | sh /opt/home/netx/backups/pa/Master-GWA-SNMP.2019-02-19.0/PA/GWA-SNMP.sh')

#pname = sys.argv[1]
#alarm_count = sys.argv[2]
#varb= "PMO-"+pname
#print "\n This is passed plugin name on commandline:",pname , "\n This will be PMO: - ",varb

v,w,Plugin_name=ReadCMT.CMTvalues()
line=ReadCMT.Class_values()
print "These values are fetched from ReadCMT module-->called function CMTvalues:",v,w,Plugin_name
print "These values are fetched from ReadCMT module-->called function Class_values:",line


varb= "PMO-"+Plugin_name
print "\n This is Plugin_Name to TEST:",Plugin_name , "\n This will be PMO: - ",varb



line1=line.strip(' ');
class_line = list(line1.split(" ")) 
class_line.pop(0)
#class_line=re.sub(r' ', "|", line1)
#class_line1 = re.sub(r'^|$',"'", class_line) 
print " This is Class List --> : ",class_line

#print "Element at 1 index:",class_line[1]

#Establishing DB connection
connstr='marocdb/marocdb@10.32.0.158:1521/h3g12cr'
conn = cx_Oracle.connect(connstr)
curs = conn.cursor()
curs.arraysize=100
####################### Putting EXCEL VALUES ############
xfile = openpyxl.load_workbook('TestReport_Format.xlsx')
sheet = xfile.get_sheet_by_name('Scorecard')
sheet['C3'] = Plugin_name 
xfile.save('TestReport_Format.xlsx')

#######################################

def Test_result(status, cellno):
    print 'This is to check the values of Status n Cellno:',status, cellno,"\n"
    #xfile = openpyxl.load_workbook('TestReport-fm_sams_enodeb_snmp_nr7.0.0.xlsx')
    xfile = openpyxl.load_workbook('TestReport_Format.xlsx')

    sheet = xfile.get_sheet_by_name('Scorecard')
    sheet[cellno] = status
    xfile.save('TestReport_Format.xlsx')
##################################################

print "\n                            Excectuing TC-001:: \n" 

named_params = {}
named_params['pn'] = varb 
#named_params['ac'] = sys.argv[2]
#curs.execute("select MO,CLASS,NAME from MO where NAME=?",varb)
curs.execute('select MO,CLASS,NAME from MO where NAME like :pn',named_params)
#print curs.fetchall() 
for column_1, column_2, column_3 in curs.fetchall():
        print column_1, "\t", column_2, "\t",column_3
        if column_3==varb:
           ID=column_1
           print '\nThis MO is going to be your MOID',ID

#print curs.bindnames() 
#curs.execute("select MOID,STRVALUE from MOATT where STRVALUE in ('fm_sams_enodeb_snmp_nr7.0.0-samsung-mib','GWA-SNMP-TRAP-162')")
curs.execute("select MO,CLASS,NAME from MO where Name like '%GWA-SNMP-FIFO%'")
for column_1, column_2, column_3 in curs.fetchall():
        print column_1, "\t", column_2, "\t",column_3 
#DATA = curs.rowcount
#print 'This is row count of the executed query:',DATA
        if column_3=='GWA-SNMP-FIFO':
                  print 'TC-001 :: PAAS'
                  Test_result( status="PASS",cellno='G30' );


print "\n                             Excectuing TC-002:: \n"


named_param = {}
named_param['ac'] = ID
named_param['od'] = "%"+v+"%"
print "Printing Dictionary value to test :",named_param['ac']
print "Printing Dictionary value to test :",named_param['od']
curs.execute("select MOID,STRVALUE from MOATT where MOID=:ac  and strvalue!='null' and strvalue like :od ",named_param)


print "MOID\tSTRVALUE\n"
for column_1, column_2 in curs.fetchall():
        print column_1, "\t", column_2
DATA = curs.rowcount
print '\n This is number of alarms as of Now in ORACLE DB for particular Plugin name:',DATA 


if DATA  == w:
   print 'TC-002 :: PAAS'
   Test_result( status="PASS",cellno='G36' );
else:
   print 'TC-002:: FAIL'
   Test_result(status ="FAIL", cellno='G36');


print "\n                         Excectuing TC-003::...... \n"
named_p = {}
named_p['ac'] = ID

curs.execute("select MOID,STRVALUE from MOATT where MOID=:ac and strvalue != 'null' and compid='-1'",named_p)

print "MOID\tSTRVALUE\n"
for column_1, column_2 in curs.fetchall():
       # print column_1, "\t", column_2
        if column_2 == 'vsaTempObjects_'+Plugin_name : 
            Test_result( status="PASS",cellno='G32' );
        if column_2 == 'PMO-'+Plugin_name :
            Test_result( status="PASS",cellno='G31' );
 
        print column_1, "\t", column_2
DATA1 = curs.rowcount
print '\n These are attributes of plugins in NetExpert:',DATA1



print "\n                         Excectuing TC-011::...... \n"

for x in class_line:

     var=x
     named_para = {}
     named_para['c']= var
     #named_para['c']= "%"+var+"%"
     #print "Printing Dictionary value to classline :",named_para['c']

     curs.execute("select CLASS,NAMINGATTRIBUTE,NAME from CLASS where NAME=:c ", named_para)


     #print "CLASS\tNAMINGATTRIBUTE\tNAME\n"

     for column_1,column_2,column_3 in curs.fetchall():
         print column_1,"\t",column_2,"\t",column_3     

#DATA1 = curs.rowcount
#print '\n This is to check the classnames in DB:',DATA1
print "\n                         BELOW IS THE LIST OF  ALL ALRMS FOR the PARTICULAR PLUGINS ...... \n"

name_p = {}
name_p['pn'] = Plugin_name
#name_p['d'] =  date 


print "ALERTID\tMOID\tMO\tCLASS_ID\tCLASS_NAME\tALARM_CREATE_DATE\tCLEAR\tALARM_DESCRITION\tAMO\tPlugin_Name\n"
curs.execute('select Distinct e.Alertid, a.MOID,b.MO,b.class AS CLASS_ID,D.NAME as CLASS_NAME,c.FIRST1,c.CLEARED,c.TEXT1,b.NAME,e.STRVALUE from MOATT a , MO b ,ALERTLOG c ,class d,Alertprop e where  a.MOID=b.MO and b.MO=c.MO and b.class=d.class and c.AlertID=e.AlertID and c.CLEARED IS null  and e.STRVALUE=:pn',name_p)

for column_1,column_2,column_3,column_4,column_5,column_6,column_7,column_8,column_9,column_10 in curs.fetchall():
         print column_1,column_2,column_3,"\t",column_4,"\t",column_5,"\t",column_6,"\t",column_7,"\t",column_8,column_9,column_10




curs.close()
conn.close()

os.rename("TestReport_Format.xlsx", "TestReport_"+Plugin_name+"."+"xlsx")
#execfile("Email.py")
#shutil.copy( "/opt/home/netx/RAJAT_AUTOMATION//CMT_Folder/TestReport_Format.xlsx", "/opt/home/netx/RAJAT_AUTOMATION/TestReport_Format.xlsx") 

